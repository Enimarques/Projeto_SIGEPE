from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Q
from django.core.files.base import ContentFile
from django.conf import settings
from .models import Visitante, Visita, Setor, VisitanteArquivado, VisitaArquivada
from .forms import VisitanteForm, VisitaForm
from .forms_departamento import SetorForm
from .misc_utils import gerar_etiqueta_pdf
import base64
import json
from django.urls import reverse
from datetime import datetime, timedelta, time
from .forms_departamento import AlterarHorarioSetorForm
from django.core.exceptions import PermissionDenied
from apps.autenticacao.decorators import admin_required, block_assessor, assessor_or_admin_required, admin_or_recepcionista_only
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from apps.veiculos.models import Veiculo
from .utils.facial_recognition_utils import get_face_embedding
from .utils.image_utils import process_image
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
import face_recognition
import numpy as np
from django.contrib.auth.models import User
from django.core.files.storage import default_storage
from django.db.models.fields.files import FieldFile
from django.core.files import File
from django.db.models import Value
from django.db.models.functions import Replace
from django.utils.timezone import get_current_timezone
import os
from django.core.cache import cache
import logging
import shutil
from pathlib import Path

# Logger do módulo
logger = logging.getLogger(__name__)

# Contexto base para todas as views do app
def get_base_context(title_suffix=''):
    return {
        'title': f'Recepção - {title_suffix}' if title_suffix else 'Recepção - URUTAU',
        'app_name': 'recepcao'
    }

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def home_recepcao(request):
    hoje = timezone.localtime().date()
    
    # Contagem de visitas
    visitas_hoje = Visita.objects.filter(data_entrada__date=hoje).count()
    visitas_em_andamento = Visita.objects.filter(data_saida__isnull=True).count()
    visitas_finalizadas_hoje = Visita.objects.filter(data_entrada__date=hoje, data_saida__isnull=False).count()
    total_visitantes = Visitante.objects.count()
    
    # Contagem de veículos
    veiculos_no_estacionamento = Veiculo.objects.filter(status='presente').count()
    total_veiculos_cadastrados = Veiculo.objects.count()
    
    # Contagem de gabinetes abertos
    gabinetes = Setor.objects.filter(tipo__in=['gabinete', 'gabinete_vereador'], ativo=True)
    gabinetes_abertos = sum(1 for g in gabinetes if g.esta_aberto())
    
    context = get_base_context()
    context.update({
        'visitas_hoje': visitas_hoje,
        'visitas_em_andamento': visitas_em_andamento,
        'visitas_finalizadas_hoje': visitas_finalizadas_hoje,
        'total_visitantes': total_visitantes,
        'veiculos_no_estacionamento': veiculos_no_estacionamento,
        'total_veiculos_cadastrados': total_veiculos_cadastrados,
        'gabinetes_abertos': gabinetes_abertos,
    })
    return render(request, 'recepcao/home_recepcao.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def lista_visitantes(request):
    # Obter parâmetros de busca
    busca = request.GET.get('busca', '')
    
    # Query base
    visitantes = Visitante.objects.all()
    
    # Aplicar filtro de busca
    if busca:
        visitantes = visitantes.filter(
            Q(nome_completo__icontains=busca) |
            Q(CPF__icontains=busca) |
            Q(email__icontains=busca)
        )
    
    # Ordenar por nome
    visitantes = visitantes.order_by('nome_completo')
    
    # Estatísticas para os cards
    visitas_em_andamento = Visita.objects.filter(data_saida__isnull=True).count()
    total_visitas = Visita.objects.count()
    
    context = get_base_context('Lista de Visitantes')
    context.update({
        'visitantes': visitantes,
        'busca': busca,
        'visitas_em_andamento': visitas_em_andamento,
        'total_visitas': total_visitas
    })
    
    return render(request, 'recepcao/lista_visitantes.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def cadastro_visitantes(request):
    if request.method == 'POST':
        form = VisitanteForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Salva o objeto na memória sem comitar no banco de dados
                visitante = form.save(commit=False)
                # O processamento da imagem, que depende de outros campos,
                # agora ocorrerá de forma segura no save() do modelo.
                visitante.save()
                form.save_m2m() # Salva as relações many-to-many, se houver
                messages.success(request, 'Visitante cadastrado com sucesso!')
                return redirect('recepcao:detalhes_visitante', pk=visitante.id)
            except Exception as e:
                # Log do erro para depuração
                logger.error("Erro crítico ao salvar visitante: %s", e)
                messages.error(request, f"Ocorreu um erro inesperado ao salvar: {e}")
        else:
            # Log dos erros de validação para depuração
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Erros de validação do formulário: %s", form.errors.as_json())
            messages.error(request, 'Por favor, corrija os erros no formulário.')
    else:
        form = VisitanteForm()
    
    context = get_base_context('Cadastro de Visitante')
    context.update({
        'form': form,
        'visitante': None,
    })
    
    return render(request, 'recepcao/cadastro_visitantes.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def detalhes_visitante(request, pk):
    visitante = get_object_or_404(Visitante, pk=pk)
    
    # Obter todas as visitas relacionadas a este visitante
    visitas = Visita.objects.filter(visitante=visitante).order_by('-data_entrada')
    
    context = get_base_context('Detalhes do Visitante')
    context.update({
        'visitante': visitante,
        'visitas': visitas
    })
    return render(request, 'recepcao/detalhes_visitante.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def editar_visitante(request, pk):
    visitante = get_object_or_404(Visitante, pk=pk)
    
    if request.method == 'POST':
        # Instancia o formulário com os dados da requisição e os arquivos enviados
        form = VisitanteForm(request.POST, request.FILES, instance=visitante)
        
        if form.is_valid():
            # Salva o objeto na memória sem comitar no banco de dados
            visitante = form.save(commit=False)
            # O processamento da imagem (incluindo a chamada para get_visitor_upload_path)
            # agora funcionará corretamente.
            visitante.save()
            form.save_m2m() # Salva as relações many-to-many, se houver
            
            # Limpa o cache da foto antiga para garantir que a nova seja exibida
            cache.delete(f'visitante_foto_{visitante.id}_thumbnail')
            cache.delete(f'visitante_foto_{visitante.id}_medium')
            cache.delete(f'visitante_foto_{visitante.id}_large')

            messages.success(request, 'Visitante atualizado com sucesso!')
            return redirect('recepcao:detalhes_visitante', pk=visitante.pk)
        else:
            # Se o formulário for inválido, exibe os erros
            error_message = "Por favor, corrija os erros abaixo. "
            for field, errors in form.errors.items():
                error_message += f" {field}: {', '.join(errors)} "
            messages.error(request, error_message)
    else:
        form = VisitanteForm(instance=visitante)
    
    context = get_base_context('Editar Visitante')
    context.update({
        'form': form,
        'visitante': visitante,
        'is_editing': True
    })
    
    return render(request, 'recepcao/cadastro_visitantes.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def registro_visitas(request):
    # Verificar se tem visitante pré-selecionado
    visitante_id = request.GET.get('visitante')
    visitante = None
    if visitante_id:
        try:
            visitante = Visitante.objects.get(id=visitante_id)
        except Visitante.DoesNotExist:
            messages.error(request, 'Visitante não encontrado.')
    
    if request.method == 'POST':
        form = VisitaForm(request.POST)
        if form.is_valid():
            try:
                visitante = Visitante.objects.get(CPF=form.cleaned_data['cpf'])
                setor = form.cleaned_data['setor']
                
                # Criar a visita
                visita = Visita.objects.create(
                    visitante=visitante,
                    setor=setor,
                    objetivo=form.cleaned_data['objetivo'],
                    observacoes=form.cleaned_data.get('observacoes', ''),
                    localizacao=setor.localizacao  # Usando a localização do setor
                )
                
                messages.success(request, 'Visita registrada com sucesso!')
                return redirect('recepcao:status_visita')
            except Visitante.DoesNotExist:
                messages.error(request, 'Visitante não encontrado.')
            except Exception as e:
                messages.error(request, f'Erro ao registrar visita: {str(e)}')
    else:
        initial_data = {}
        if visitante:
            initial_data['cpf'] = visitante.CPF
        form = VisitaForm(initial=initial_data)
    
    context = get_base_context('Registro de Visitas')
    context.update({
        'form': form,
        'visitante_pre_selecionado': visitante
    })
    
    return render(request, 'recepcao/registro_visitas.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def buscar_visitante(request):
    query = request.GET.get('query', None)
    
    if not query or len(query) < 3:
        return JsonResponse({'success': False, 'message': 'Digite ao menos 3 caracteres para buscar.'})
    
    # Busca simples e direta
    visitantes = Visitante.objects.filter(
        Q(nome_completo__icontains=query) |
        Q(CPF__icontains=query)
    ).order_by('nome_completo')[:10]

    if not visitantes.exists():
        return JsonResponse({
            'success': False,
            'message': 'Nenhum visitante encontrado.'
        })

    # Prepara os dados para a resposta JSON
    visitantes_data = []
    for v in visitantes:
        # Buscar a última visita do visitante
        ultima_visita = Visita.objects.filter(visitante=v).order_by('-data_entrada').first()
        
        visitante_data = {
            'id': v.id,
            'nome_completo': v.nome_completo,
            'cpf': v.CPF, # Retorna o CPF formatado
            'telefone': v.telefone,
            'foto': v.foto.url if v.foto else None,
            'ultima_visita': ultima_visita.data_entrada.strftime('%d/%m/%Y') if ultima_visita else None,
            'total_visitas': Visita.objects.filter(visitante=v).count(),
            'email': v.email if v.email else 'Não informado'
        }
        visitantes_data.append(visitante_data)

    return JsonResponse({'success': True, 'visitantes': visitantes_data})

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def buscar_setores(request):
    tipo = request.GET.get('tipo', 'departamento')
    
    # Filtrar setores por tipo
    setores = Setor.objects.filter(tipo=tipo).order_by('id')
    
    if not setores.exists():
        return JsonResponse({
            'success': False,
            'message': f'Nenhum {tipo} encontrado.'
        })
    
    # Preparar dados dos setores
    setores_data = []
    for setor in setores:
        dados = {
            'id': setor.id,
            'nome': setor.nome_vereador if tipo in ['gabinete', 'gabinete_vereador'] else setor.nome_local,
            'nome_vereador': setor.nome_vereador if tipo in ['gabinete', 'gabinete_vereador'] else None,
            'nome_local': setor.nome_local if tipo == 'departamento' else None,
            'tipo': setor.tipo,
            'localizacao': setor.get_localizacao_display(),
            'funcao': setor.get_funcao_display() if setor.funcao else None,
            'email': setor.email_vereador if tipo in ['gabinete', 'gabinete_vereador'] else setor.email,
            'foto_url': None, # Inicialmente nulo
            'aberto_agora': setor.esta_aberto()
        }
        
        # Prioridade: foto do gabinete (vereador), depois foto do assessor
        if tipo in ['gabinete', 'gabinete_vereador'] and setor.foto:
            dados['foto_url'] = setor.foto.url
        elif setor.usuario and hasattr(setor.usuario, 'assessor') and setor.usuario.assessor.foto:
            dados['foto_url'] = setor.usuario.assessor.foto.url

        setores_data.append(dados)
    
    return JsonResponse({
        'success': True,
        'setores': setores_data
    })

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def historico_visitas(request):
    # Filtros
    status = request.GET.get('status')
    periodo = request.GET.get('periodo')
    busca = request.GET.get('busca')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    # Query base
    visitas = Visita.objects.all()
    if logger.isEnabledFor(logging.DEBUG):
        logger.debug("Total de visitas antes dos filtros: %s", visitas.count())
    
    # Aplicar filtros
    if status:
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Aplicando filtro de status: %s", status)
        if status == 'em_andamento':
            visitas = visitas.filter(data_saida__isnull=True)
        elif status == 'finalizada':
            visitas = visitas.filter(data_saida__isnull=False)
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Total após filtro de status: %s", visitas.count())
    
    # Filtro de data específica (tem prioridade sobre período rápido)
    if data_inicio or data_fim:
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Aplicando filtro de data específica: %s a %s", data_inicio, data_fim)
        try:
            if data_inicio:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                visitas = visitas.filter(data_entrada__date__gte=data_inicio_obj)
            if data_fim:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
                # Adicionar 1 dia e subtrair 1 segundo para incluir o dia inteiro
                data_fim_completa = datetime.combine(data_fim_obj, time.max)
                visitas = visitas.filter(data_entrada__lte=data_fim_completa)
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Total após filtro de data específica: %s", visitas.count())
        except ValueError as e:
            logger.warning(f"Erro ao processar filtro de data: {e}")
    elif periodo:
        # Filtro de período rápido (só aplica se não houver data específica)
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Aplicando filtro de período: %s", periodo)
        hoje = timezone.localtime().date()
        if periodo == 'hoje':
            visitas = visitas.filter(data_entrada__date=hoje)
        elif periodo == 'semana':
            inicio_semana = hoje - timedelta(days=hoje.weekday())
            visitas = visitas.filter(data_entrada__date__gte=inicio_semana)
        elif periodo == 'mes':
            inicio_mes = hoje.replace(day=1)
            visitas = visitas.filter(data_entrada__date__gte=inicio_mes)
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Total após filtro de período: %s", visitas.count())
    
    if busca:
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Aplicando busca: %s", busca)
        visitas = visitas.filter(
            Q(visitante__nome_completo__icontains=busca) |
            Q(visitante__CPF__icontains=busca)
        )
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Total após busca: %s", visitas.count())
    
    # Ordenação
    visitas = visitas.order_by('-data_entrada')
    
    # Paginação
    paginator = Paginator(visitas, 10)  # 10 visitas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_visitas = visitas.count()
    visitas_em_andamento = visitas.filter(data_saida__isnull=True).count()
    visitas_finalizadas = visitas.filter(data_saida__isnull=False).count()
    
    context = get_base_context('Histórico de Visitas')
    context.update({
        'page_obj': page_obj,
        'total_visitas': total_visitas,
        'visitas_em_andamento': visitas_em_andamento,
        'visitas_finalizadas': visitas_finalizadas,
        # Manter filtros selecionados
        'status_filtro': status,
        'periodo_filtro': periodo,
        'busca': busca,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    })
    
    return render(request, 'recepcao/historico_visitas.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def status_visita(request):
    """
    Exibe as visitas em andamento com filtros.
    """
    # Obter parâmetros do filtro (manter como string para preservar no template)
    data_str = request.GET.get('data', '')
    hora_inicio_str = request.GET.get('hora_inicio', '')
    hora_fim_str = request.GET.get('hora_fim', '')
    localizacao = request.GET.get('localizacao', '')
    setor = request.GET.get('setor', '')
    busca_nome = request.GET.get('busca_nome', '')

    # Iniciar queryset com visitas não finalizadas e status em andamento
    visitas = Visita.objects.filter(
        data_saida__isnull=True,
        status='em_andamento'
    )

    # Aplicar filtros
    if data_str:
        try:
            data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
            visitas = visitas.filter(data_entrada__date=data_obj)
        except (ValueError, TypeError):
            messages.error(request, 'Data inválida')
            data_str = ''  # Limpar se inválida

    if hora_inicio_str:
        try:
            hora_inicio_obj = datetime.strptime(hora_inicio_str, '%H:%M').time()
            visitas = visitas.filter(data_entrada__time__gte=hora_inicio_obj)
        except (ValueError, TypeError):
            messages.error(request, 'Hora de início inválida')
            hora_inicio_str = ''  # Limpar se inválida

    if hora_fim_str:
        try:
            hora_fim_obj = datetime.strptime(hora_fim_str, '%H:%M').time()
            visitas = visitas.filter(data_entrada__time__lte=hora_fim_obj)
        except (ValueError, TypeError):
            messages.error(request, 'Hora de fim inválida')
            hora_fim_str = ''  # Limpar se inválida

    if localizacao:
        visitas = visitas.filter(localizacao=localizacao)

    if setor:
        visitas = visitas.filter(setor_id=setor)

    if busca_nome:
        visitas = visitas.filter(
            Q(visitante__nome_completo__icontains=busca_nome) |
            Q(visitante__nome_social__icontains=busca_nome)
        )

    # Ordenar por data de entrada mais recente
    visitas = visitas.order_by('-data_entrada')

    # Estatísticas principais para os cards conforme modelo de referência
    total_em_andamento = visitas.count()
    total_visitantes = Visitante.objects.count()
    
    # Métricas de veículos
    veiculos_no_estacionamento = Veiculo.objects.filter(
        status='presente'
    ).count()
    total_veiculos_cadastrados = Veiculo.objects.count()
    
    # Estatísticas por localização (para gráficos ou outras visualizações se necessário)
    total_por_local = {
        'terreo': visitas.filter(localizacao='terreo').count(),
        'plenario': visitas.filter(localizacao='plenario').count(),
        'primeiro_piso': visitas.filter(localizacao='primeiro_piso').count(),
        'segundo_piso': visitas.filter(localizacao='segundo_piso').count()
    }

    context = get_base_context('Status das Visitas')
    context.update({
        'visitas': visitas,
        'total_em_andamento': total_em_andamento,
        'total_visitantes': total_visitantes,
        'veiculos_no_estacionamento': veiculos_no_estacionamento,
        'total_veiculos_cadastrados': total_veiculos_cadastrados,
        'total_por_local': total_por_local,
        'setores': Setor.objects.filter(ativo=True).order_by('id'),
        'localizacoes': dict(Visita.LOCALIZACAO_CHOICES),
        # Manter filtros selecionados como strings originais
        'data_filtro': data_str,
        'hora_inicio_filtro': hora_inicio_str,
        'hora_fim_filtro': hora_fim_str,
        'localizacao_filtro': localizacao,
        'setor_filtro': setor,
        'busca_nome_filtro': busca_nome
    })

    return render(request, 'recepcao/status_visita.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def finalizar_visita(request, visita_id):
    """
    Finaliza uma visita em andamento.
    """
    try:
        visita = Visita.objects.get(id=visita_id)
    except Visita.DoesNotExist:
        messages.error(request, 'Visita não encontrada. Ela pode ter sido finalizada ou excluída por outro usuário.')
        return redirect('recepcao:lista_visitantes')

    # Se já finalizada, informa e volta para os detalhes do visitante
    if visita.data_saida or visita.status != 'em_andamento':
        messages.error(request, 'Esta visita já foi finalizada.')
        return redirect('recepcao:detalhes_visitante', pk=visita.visitante.id)

    try:
        # Finaliza a visita
        visita.data_saida = timezone.now()
        visita.status = 'finalizada'
        visita.save()

        messages.success(request, 'Visita finalizada com sucesso!')
        return redirect('recepcao:detalhes_visitante', pk=visita.visitante.id)
    except Exception as e:
        messages.error(request, f'Erro ao finalizar visita: {str(e)}')
        return redirect('recepcao:detalhes_visitante', pk=visita.visitante.id)

@login_required(login_url='autenticacao:login_sistema')
@admin_required
def excluir_visita(request, pk):
    visita = get_object_or_404(Visita, pk=pk)
    logging.getLogger('audit').info(
        f"user={request.user.username} action=DELETE_VISITA visita_id={visita.id} visitante_id={visita.visitante.id} ip={request.META.get('REMOTE_ADDR')}"
    )
    visita.delete()
    messages.success(request, 'Visita excluída com sucesso!')
    return redirect('recepcao:historico_visitas')

@login_required(login_url='autenticacao:login_sistema')
def excluir_setor(request, pk):
    # Verificar se o usuário é superadmin
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem excluir setores.')
        return redirect('admin:recepcao_setor_changelist')
        
    setor = get_object_or_404(Setor, pk=pk)
    
    if request.method == 'POST':
        try:
            # Excluir todas as visitas relacionadas a este setor
            visitas_relacionadas = Visita.objects.filter(setor=setor)
            qtd_visitas = visitas_relacionadas.count()
            
            # Excluir as visitas primeiro
            if qtd_visitas > 0:
                visitas_relacionadas.delete()
                messages.warning(request, f'{qtd_visitas} visita(s) relacionada(s) foram excluída(s).')
            
            # Verificar se o setor está relacionado a um gabinete
            try:
                from apps.gabinetes.models import Gabinete
                gabinetes = Gabinete.objects.filter(setor=setor)
                if gabinetes.exists():
                    # Excluir os gabinetes relacionados
                    qtd_gabinetes = gabinetes.count()
                    gabinetes.delete()
                    messages.warning(request, f'{qtd_gabinetes} gabinete(s) relacionado(s) foram excluído(s).')
            except ImportError:
                pass
                
            # Verificar se existem assessores relacionados ao setor
            assessores = Assessor.objects.filter(
                ativo=True,
                departamento__isnull=False
            ).order_by('id')
            if assessores.exists():
                # Desassociar os assessores deste setor
                qtd_assessores = assessores.count()
                assessores.update(departamento=None)
                messages.warning(request, f'{qtd_assessores} assessor(es) relacionado(s) foram desassociados.')
                
            # Depois excluir o setor
            nome_setor = setor.nome_vereador if setor.tipo in ['gabinete', 'gabinete_vereador'] else setor.nome_local
            logging.getLogger('audit').info(
                f"user={request.user.username} action=DELETE_SETOR setor_id={setor.id} nome='{nome_setor}' ip={request.META.get('REMOTE_ADDR')}"
            )
            setor.delete()
            messages.success(request, f'Setor "{nome_setor}" excluído com sucesso!')
            
            # Redirecionar para a lista de setores no admin
            return redirect('admin:recepcao_setor_changelist')
        except Exception as e:
            messages.error(request, f"Erro ao excluir setor: {str(e)}")
            return redirect('admin:recepcao_setor_changelist')
    
    # Mostra um template de confirmação
    context = {
        'setor': setor,
        'visitas_relacionadas': Visita.objects.filter(setor=setor),
        'objeto_tipo': 'Setor',
        'objeto_nome': setor.nome_vereador if setor.tipo in ['gabinete', 'gabinete_vereador'] else setor.nome_local
    }
    
    # Verificar se o setor está relacionado a um gabinete
    try:
        from apps.gabinetes.models import Gabinete
        context['gabinetes_relacionados'] = Gabinete.objects.filter(setor=setor)
    except ImportError:
        context['gabinetes_relacionados'] = []
    
    # Verificar se existem assessores relacionados ao setor
    context['assessores_relacionados'] = assessores
    
    return render(request, 'recepcao/confirmar_exclusao_setor.html', context)

def arquivar_visitante(visitante, usuario_arquivou):
    """
    Arquivar visitante e suas visitas relacionadas.
    Move os dados para as tabelas de arquivo ao invés de deletar.
    """
    try:
        # Verificar se o visitante já foi arquivado (proteção contra duplicação)
        visitante_arquivado_existente = VisitanteArquivado.objects.filter(
            id_original=visitante.id
        ).first()
        
        if visitante_arquivado_existente:
            logger.warning(
                f"Visitante {visitante.id} ({visitante.nome_completo}) já foi arquivado anteriormente "
                f"(ID arquivado: {visitante_arquivado_existente.id}). Retornando registro existente."
            )
            # Retornar o registro existente e contar as visitas já arquivadas
            num_visitas = visitante_arquivado_existente.visitas_arquivadas.count()
            return visitante_arquivado_existente, num_visitas
        
        # Criar visitante arquivado
        visitante_arquivado = VisitanteArquivado(
            nome_completo=visitante.nome_completo,
            nome_social=visitante.nome_social,
            data_nascimento=visitante.data_nascimento,
            CPF=visitante.CPF,
            telefone=visitante.telefone,
            email=visitante.email,
            estado=visitante.estado,
            cidade=visitante.cidade,
            bairro=visitante.bairro,
            logradouro=visitante.logradouro,
            numero=visitante.numero,
            complemento=visitante.complemento,
            CEP=visitante.CEP,
            biometric_vector=visitante.biometric_vector,
            id_original=visitante.id,
            data_cadastro_original=visitante.data_cadastro,
            usuario_arquivou=usuario_arquivou
        )
        
        # Copiar arquivos de foto para a pasta de arquivados
        foto_fields = ['foto', 'foto_thumbnail', 'foto_medium', 'foto_large']
        for field_name in foto_fields:
            original_field = getattr(visitante, field_name, None)
            if original_field and original_field.name:
                try:
                    # Caminho original
                    original_path = Path(original_field.path)
                    if original_path.exists():
                        # Criar caminho de destino
                        filename = original_path.name
                        dest_dir = Path(settings.MEDIA_ROOT) / 'arquivados' / 'fotos_visitantes'
                        dest_dir.mkdir(parents=True, exist_ok=True)
                        dest_path = dest_dir / filename
                        
                        # Copiar arquivo
                        shutil.copy2(original_path, dest_path)
                        
                        # Atribuir ao campo do visitante arquivado
                        relative_path = f'arquivados/fotos_visitantes/{filename}'
                        setattr(visitante_arquivado, field_name, relative_path)
                except Exception as e:
                    logger.warning(f"Erro ao copiar {field_name} do visitante {visitante.id}: {e}")
        
        visitante_arquivado.save()
        
        # Arquivar visitas relacionadas
        visitas_relacionadas = Visita.objects.filter(visitante=visitante)
        num_visitas = 0
        visitas_ids = []
        
        for visita in visitas_relacionadas:
            visita_arquivada = VisitaArquivada(
                visitante_arquivado=visitante_arquivado,
                id_original=visita.id,
                nome_setor=str(visita.setor),
                localizacao=visita.localizacao,
                objetivo=visita.objetivo,
                observacoes=visita.observacoes,
                data_entrada=visita.data_entrada,
                data_saida=visita.data_saida,
                status=visita.status
            )
            visita_arquivada.save()
            visitas_ids.append(visita.id)
            num_visitas += 1
        
        # Deletar as visitas originais primeiro (antes de deletar o visitante)
        # Isso é necessário porque Visita tem on_delete=models.PROTECT
        if visitas_ids:
            Visita.objects.filter(id__in=visitas_ids).delete()
        
        # Agora deletar o visitante original (que vai deletar as fotos originais)
        visitante.delete()
        
        return visitante_arquivado, num_visitas
        
    except Exception as e:
        logger.error(f"Erro ao arquivar visitante {visitante.id}: {e}", exc_info=True)
        raise


@login_required(login_url='autenticacao:login_sistema')
@admin_required
def excluir_visitante(request, pk):
    # Verificar se o visitante já foi arquivado antes de buscar
    visitante_arquivado_existente = VisitanteArquivado.objects.filter(id_original=pk).first()
    if visitante_arquivado_existente:
        messages.info(
            request,
            f'Este visitante já foi arquivado anteriormente em {visitante_arquivado_existente.data_arquivamento.strftime("%d/%m/%Y %H:%M")}. '
            f'Você pode visualizá-lo na lista de visitantes arquivados.'
        )
        return redirect('recepcao:lista_visitantes')
    
    visitante = get_object_or_404(Visitante, pk=pk)
    visitas_relacionadas = Visita.objects.filter(visitante=visitante)
    
    if request.method == 'POST':
        try:
            # Verificar novamente se não foi arquivado entre a requisição GET e POST
            # (proteção contra múltiplas submissões)
            if not Visitante.objects.filter(pk=pk).exists():
                messages.warning(request, 'Este visitante já foi arquivado ou não existe mais.')
                return redirect('recepcao:lista_visitantes')
            
            # Arquivar visitante ao invés de deletar
            nome_visitante = visitante.nome_completo
            visitante_arquivado, num_visitas = arquivar_visitante(visitante, request.user.username)
            
            logging.getLogger('audit').info(
                f"user={request.user.username} action=ARCHIVE_VISITANTE visitante_id={visitante_arquivado.id_original} "
                f"nome='{nome_visitante}' visitas_arquivadas={num_visitas} arquivado_id={visitante_arquivado.id} "
                f"ip={request.META.get('REMOTE_ADDR')}"
            )
            
            messages.success(
                request, 
                f'Visitante "{nome_visitante}" e seu histórico de {num_visitas} visita(s) foram arquivados com sucesso! '
                f'Os dados serão mantidos por 6 meses antes da exclusão definitiva.'
            )
            return redirect('recepcao:lista_visitantes')
        except Exception as e:
            logger.error(f"Erro ao arquivar visitante: {e}", exc_info=True)
            messages.error(request, f'Erro ao arquivar visitante: {str(e)}')
            return redirect('recepcao:detalhes_visitante', pk=pk)
    
    context = get_base_context('Excluir Visitante')
    context.update({
        'visitante': visitante,
        'visitas_relacionadas': visitas_relacionadas
    })
    return render(request, 'recepcao/confirmar_exclusao_visitante.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_required
def lista_visitantes_arquivados(request):
    """Lista visitantes arquivados - apenas para administradores"""
    busca = request.GET.get('busca', '')
    
    visitantes_arquivados = VisitanteArquivado.objects.all()
    
    if busca:
        visitantes_arquivados = visitantes_arquivados.filter(
            Q(nome_completo__icontains=busca) |
            Q(CPF__icontains=busca) |
            Q(email__icontains=busca)
        )
    
    visitantes_arquivados = visitantes_arquivados.order_by('-data_arquivamento')
    
    # Paginação
    paginator = Paginator(visitantes_arquivados, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_arquivados = VisitanteArquivado.objects.count()
    from datetime import timedelta
    from django.utils import timezone
    cutoff = timezone.now() - timedelta(days=180)
    proximos_a_expirar = VisitanteArquivado.objects.filter(data_arquivamento__lte=cutoff).count()
    
    context = get_base_context('Visitantes Arquivados')
    context.update({
        'page_obj': page_obj,
        'busca': busca,
        'total_arquivados': total_arquivados,
        'proximos_a_expirar': proximos_a_expirar,
    })
    
    return render(request, 'recepcao/lista_visitantes_arquivados.html', context)


@login_required(login_url='autenticacao:login_sistema')
@admin_required
def detalhes_visitante_arquivado(request, pk):
    """Detalhes de um visitante arquivado - apenas para administradores"""
    visitante_arquivado = get_object_or_404(VisitanteArquivado, pk=pk)
    
    # Obter todas as visitas arquivadas relacionadas
    visitas_arquivadas = VisitaArquivada.objects.filter(
        visitante_arquivado=visitante_arquivado
    ).order_by('-data_entrada')
    
    # Calcular dias até expiração
    from datetime import timedelta
    from django.utils import timezone
    data_expiracao = visitante_arquivado.data_arquivamento + timedelta(days=180)
    dias_restantes = (data_expiracao - timezone.now()).days
    
    context = get_base_context('Detalhes do Visitante Arquivado')
    context.update({
        'visitante_arquivado': visitante_arquivado,
        'visitas_arquivadas': visitas_arquivadas,
        'data_expiracao': data_expiracao,
        'dias_restantes': dias_restantes,
    })
    
    return render(request, 'recepcao/detalhes_visitante_arquivado.html', context)


@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def gerar_etiqueta(request, visita_id):
    visita = get_object_or_404(Visita, pk=visita_id)
    context = get_base_context('Etiqueta da Visita')
    context.update({
        'visita': visita,
        'nome_exibicao': visita.visitante.nome_social if visita.visitante.nome_social else visita.visitante.nome_completo
    })
    return render(request, 'recepcao/etiqueta_visita.html', context)

@login_required(login_url='autenticacao:login_sistema')
@assessor_or_admin_required
def alterar_horario_departamento(request):
    # Verificar se o usuário é um assessor
    try:
        setor = request.user.setor_responsavel
        if not setor:
            messages.error(request, 'Você não tem permissão para alterar horários de departamentos.')
            return redirect('recepcao:home_recepcao')
    except:
        messages.error(request, 'Você não tem permissão para alterar horários de departamentos.')
        return redirect('recepcao:home_recepcao')
    
    # Obter o departamento do assessor
    departamento = setor
    
    if request.method == 'POST':
        form = AlterarHorarioSetorForm(request.POST, instance=departamento, assessor=setor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horário do departamento alterado com sucesso!')
            return redirect('recepcao:home_recepcao')
    else:
        form = AlterarHorarioSetorForm(instance=departamento, assessor=setor)
    
    context = get_base_context('Alterar Horário do Departamento')
    context.update({
        'form': form,
        'departamento': departamento
    })
    
    return render(request, 'recepcao/alterar_horario_departamento.html', context)

@login_required(login_url='autenticacao:login_sistema')
@assessor_or_admin_required
@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def home_gabinetes(request):
    """View para a página inicial dos gabinetes."""
    # Se for assessor, mostra só o gabinete dele
    if hasattr(request.user, 'setor_responsavel') and request.user.setor_responsavel and request.user.setor_responsavel.tipo in ['gabinete', 'gabinete_vereador']:
        gabinetes = Setor.objects.filter(id=request.user.setor_responsavel.id, tipo__in=['gabinete', 'gabinete_vereador'])
    else:
        gabinetes = Setor.objects.filter(tipo__in=['gabinete', 'gabinete_vereador'])

    # Calcula estatísticas
    total_visitas = Visita.objects.count()
    visitas_hoje = Visita.objects.filter(
        data_entrada__date=timezone.now().date()
    ).count()
    visitas_em_andamento = Visita.objects.filter(
        data_saida__isnull=True
    ).count()
    hoje = timezone.now()
    primeiro_dia_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    visitas_mes = Visita.objects.filter(
        data_entrada__gte=primeiro_dia_mes
    ).count()

    context = {
        'gabinetes': gabinetes,
        'total_visitas': total_visitas,
        'visitas_hoje': visitas_hoje,
        'visitas_em_andamento': visitas_em_andamento,
        'visitas_mes': visitas_mes,
    }
    return render(request, 'recepcao/home_gabinetes.html', context)

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def detalhes_gabinete(request, gabinete_id):
    gabinete = get_object_or_404(Setor, id=gabinete_id, tipo__in=['gabinete', 'gabinete_vereador'])
    
    # Obter filtros da requisição
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    objetivo = request.GET.get('objetivo')
    formato = request.GET.get('formato')
    
    # Query base
    visitas = Visita.objects.filter(setor=gabinete)
    
    # Aplicar filtros
    if data_inicio and data_fim:
        visitas = visitas.filter(
            data_entrada__date__range=[data_inicio, data_fim]
        )
    if status:
        visitas = visitas.filter(status=status)
    if objetivo:
        visitas = visitas.filter(objetivo=objetivo)
    
    # Ordenar por data de entrada (mais recentes primeiro)
    visitas = visitas.order_by('-data_entrada')
    
    # Verificar se é uma requisição de exportação
    if formato in ['pdf', 'excel']:
        if formato == 'pdf':
            return gerar_pdf_visitas(visitas, gabinete)
        else:
            return gerar_excel_visitas(visitas, gabinete)
    
    # Paginação
    paginator = Paginator(visitas, 10)  # 10 itens por página
    page = request.GET.get('page')
    visitas_paginadas = paginator.get_page(page)
    
    # Estatísticas
    visitas_hoje = visitas.filter(data_entrada__date=timezone.now().date()).count()
    visitas_em_andamento = visitas.filter(status='em_andamento').count()
    total_visitas = visitas.count()
    visitas_mes = visitas.filter(data_entrada__month=timezone.now().month).count()
    
    context = get_base_context(f'Detalhes do Gabinete - {gabinete.nome_vereador}')
    context.update({
        'gabinete': gabinete,
        'visitas': visitas_paginadas,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'status': status,
            'objetivo': objetivo
        },
        'visitas_hoje': visitas_hoje,
        'visitas_em_andamento': visitas_em_andamento,
        'total_visitas': total_visitas,
        'visitas_mes': visitas_mes,
        'TIPOS_VISITA': Visita.OBJETIVO_CHOICES
    })
    
    return render(request, 'recepcao/detalhes_gabinete.html', context)

@login_required(login_url='autenticacao:login_sistema')
def editar_gabinete(request, gabinete_id):
    gabinete = get_object_or_404(Setor, id=gabinete_id, tipo__in=['gabinete', 'gabinete_vereador'])
    # Só o assessor responsável pode editar
    if not hasattr(request.user, 'setor_responsavel') or request.user.setor_responsavel.id != gabinete.id:
        raise PermissionDenied('Você não tem permissão para editar este gabinete.')

    if request.method == 'POST':
        form = SetorForm(request.POST, instance=gabinete, hide_responsavel=True)
        if form.is_valid():
            form.save()
            messages.success(request, 'Informações do gabinete atualizadas com sucesso!')
            return redirect('recepcao:detalhes_gabinete', gabinete_id=gabinete.id)
    else:
        form = SetorForm(instance=gabinete, hide_responsavel=True)

    context = get_base_context('Editar Gabinete')
    context.update({
        'form': form,
        'gabinete': gabinete,
        'is_editing': True
    })
    return render(request, 'recepcao/editar_gabinete.html', context)

def gerar_pdf_visitas(visitas, gabinete):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="visitas_{gabinete.nome_vereador}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    # Criar o PDF
    p = canvas.Canvas(response)
    
    # Configurações iniciais
    p.setTitle(f"Relatório de Visitas - {gabinete.nome_vereador}")
    p.setFont("Helvetica-Bold", 16)
    
    # Título
    p.drawString(50, 800, f"Relatório de Visitas - {gabinete.nome_vereador}")
    p.setFont("Helvetica", 12)
    p.drawString(50, 780, f"Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Cabeçalho da tabela
    p.setFont("Helvetica-Bold", 10)
    headers = ["Visitante", "Data/Hora Entrada", "Data/Hora Saída", "Status", "Tipo"]
    col_widths = [150, 100, 100, 80, 100]  # Ajustado para melhor distribuição
    col_positions = [50, 200, 300, 400, 480]  # Posições ajustadas das colunas
    y = 750
    
    # Desenhar cabeçalhos
    for i, header in enumerate(headers):
        p.drawString(col_positions[i], y, header)
    
    # Dados da tabela
    p.setFont("Helvetica", 9)  # Fonte um pouco menor para melhor ajuste
    y -= 20
    
    for visita in visitas:
        if y < 50:  # Nova página se necessário
            p.showPage()
            y = 750
            # Redesenhar cabeçalho
            p.setFont("Helvetica-Bold", 10)
            for i, header in enumerate(headers):
                p.drawString(col_positions[i], y, header)
            p.setFont("Helvetica", 9)
            y -= 20
        
        # Visitante
        p.drawString(col_positions[0], y, visita.visitante.nome_completo[:30] + "..." if len(visita.visitante.nome_completo) > 30 else visita.visitante.nome_completo)
        
        # Data/Hora Entrada
        p.drawString(col_positions[1], y, visita.data_entrada.strftime("%d/%m/%Y %H:%M"))
        
        # Data/Hora Saída
        p.drawString(col_positions[2], y, visita.data_saida.strftime("%d/%m/%Y %H:%M") if visita.data_saida else "-")
        
        # Status
        p.drawString(col_positions[3], y, visita.get_status_display())
        
        # Tipo
        p.drawString(col_positions[4], y, visita.get_objetivo_display())
        
        y -= 20
    
    p.save()
    return response

def gerar_excel_visitas(visitas, gabinete):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="visitas_{gabinete.nome_vereador}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    # Criar o arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitas"
    
    # Cabeçalho
    headers = ["Visitante", "Data/Hora Entrada", "Data/Hora Saída", "Status", "Tipo"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dados
    for row, visita in enumerate(visitas, 2):
        ws.cell(row=row, column=1, value=visita.visitante.nome_completo)
        ws.cell(row=row, column=2, value=visita.data_entrada.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=3, value=visita.data_saida.strftime("%d/%m/%Y %H:%M") if visita.data_saida else "-")
        ws.cell(row=row, column=4, value=visita.get_status_display())
        ws.cell(row=row, column=5, value=visita.get_objetivo_display())
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    wb.save(response)
    return response

def exportar_relatorio(visitas, formato):
    if formato == 'pdf':
        return gerar_pdf(visitas)
    else:
        return gerar_excel(visitas)

def gerar_pdf(visitas):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from io import BytesIO
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # Cabeçalho
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 750, "Relatório de Visitas")
    
    # Tabela
    y = 700
    p.setFont("Helvetica", 10)
    for visita in visitas:
        p.drawString(50, y, f"Visitante: {visita.visitante.nome_completo}")
        p.drawString(50, y-20, f"Data Entrada: {visita.data_entrada.strftime('%d/%m/%Y %H:%M')}")
        p.drawString(50, y-40, f"Status: {visita.get_status_display()}")
        y -= 60
        
        if y < 50:
            p.showPage()
            y = 750
    
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

def gerar_excel(visitas):
    from openpyxl import Workbook
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitas"
    
    # Cabeçalhos
    ws.append(['Visitante', 'Data Entrada', 'Data Saída', 'Status', 'Tipo'])
    
    # Dados
    for visita in visitas:
        ws.append([
            visita.visitante.nome_completo,
            visita.data_entrada.strftime('%d/%m/%Y %H:%M'),
            visita.data_saida.strftime('%d/%m/%Y %H:%M') if visita.data_saida else '-',
            visita.get_status_display(),
            visita.get_tipo_visita_display()
        ])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@login_required(login_url='autenticacao:login_sistema')
@block_assessor
def home_departamentos(request):
    departamentos = Setor.objects.filter(tipo='departamento')
    context = {
        'departamentos': departamentos,
    }
    return render(request, 'recepcao/home_departamentos.html', context)

@login_required(login_url='autenticacao:login_sistema')
def detalhes_departamento(request, departamento_id):
    departamento = get_object_or_404(Setor, id=departamento_id, tipo='departamento')
    visitas = Visita.objects.filter(setor=departamento).order_by('-data_entrada')
    context = {
        'departamento': departamento,
        'visitas': visitas,
    }
    return render(request, 'recepcao/detalhes_departamento.html', context)

@login_required(login_url='autenticacao:login_sistema')
def visitas_tabela_departamento(request, departamento_id):
    departamento = get_object_or_404(Setor, id=departamento_id, tipo='departamento')
    visitas = Visita.objects.filter(setor=departamento).order_by('-data_entrada')
    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    objetivo = request.GET.get('objetivo')
    if data_inicio:
        visitas = visitas.filter(data_entrada__date__gte=data_inicio)
    if data_fim:
        visitas = visitas.filter(data_entrada__date__lte=data_fim)
    if status:
        visitas = visitas.filter(status=status)
    if objetivo:
        visitas = visitas.filter(objetivo=objetivo)
    html = render_to_string('recepcao/includes/tabela_visitas_departamento.html', {'visitas': visitas})
    return JsonResponse({'html': html})

@login_required(login_url='autenticacao:login_sistema')
def visitas_tabela_gabinete(request, gabinete_id):
    gabinete = get_object_or_404(Setor, id=gabinete_id, tipo__in=['gabinete', 'gabinete_vereador'])
    visitas = Visita.objects.filter(setor=gabinete).order_by('-data_entrada')
    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    objetivo = request.GET.get('objetivo')
    if data_inicio:
        visitas = visitas.filter(data_entrada__date__gte=data_inicio)
    if data_fim:
        visitas = visitas.filter(data_entrada__date__lte=data_fim)
    if status:
        visitas = visitas.filter(status=status)
    if objetivo:
        visitas = visitas.filter(objetivo=objetivo)
    html = render_to_string('recepcao/includes/tabela_visitas_gabinete.html', {'visitas': visitas})
    return JsonResponse({'html': html})



@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def status_visita_ajax(request):
    """
    Retorna apenas a tabela de visitas filtrada via AJAX.
    """
    # Usar a mesma lógica da view principal
    data = request.GET.get('data')
    hora_inicio = request.GET.get('hora_inicio')
    hora_fim = request.GET.get('hora_fim')
    localizacao = request.GET.get('localizacao')
    setor = request.GET.get('setor')
    busca_nome = request.GET.get('busca_nome')

    # Iniciar queryset com visitas não finalizadas e status em andamento
    visitas = Visita.objects.filter(
        data_saida__isnull=True,
        status='em_andamento'
    )

    # Aplicar filtros
    if data:
        try:
            data = datetime.strptime(data, '%Y-%m-%d').date()
            visitas = visitas.filter(data_entrada__date=data)
        except (ValueError, TypeError):
            pass

    if hora_inicio:
        try:
            hora_inicio = datetime.strptime(hora_inicio, '%H:%M').time()
            visitas = visitas.filter(data_entrada__time__gte=hora_inicio)
        except (ValueError, TypeError):
            pass

    if hora_fim:
        try:
            hora_fim = datetime.strptime(hora_fim, '%H:%M').time()
            visitas = visitas.filter(data_entrada__time__lte=hora_fim)
        except (ValueError, TypeError):
            pass

    if localizacao:
        visitas = visitas.filter(localizacao=localizacao)

    if setor:
        visitas = visitas.filter(setor_id=setor)

    if busca_nome:
        from django.db.models import Q
        visitas = visitas.filter(
            Q(visitante__nome_completo__icontains=busca_nome) |
            Q(visitante__nome_social__icontains=busca_nome)
        )

    # Ordenar por data de entrada mais recente
    visitas = visitas.order_by('-data_entrada')

    # Renderizar apenas a tabela
    table_html = render_to_string(
        'recepcao/includes/tabela_visitas_status.html',
        {'visitas': visitas},
        request=request
    )

    return JsonResponse({
        'success': True,
        'html': table_html,
        'total_visitas': visitas.count()
    })

@login_required(login_url='autenticacao:login_sistema')
@admin_or_recepcionista_only
def historico_visitas_ajax(request):
    """
    Retorna apenas a tabela do histórico de visitas filtrada via AJAX.
    """
    # Usar a mesma lógica da view principal
    status = request.GET.get('status')
    periodo = request.GET.get('periodo')
    busca = request.GET.get('busca')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    # Query base
    visitas = Visita.objects.all()
    
    # Aplicar filtros
    if status:
        if status == 'em_andamento':
            visitas = visitas.filter(data_saida__isnull=True)
        elif status == 'finalizada':
            visitas = visitas.filter(data_saida__isnull=False)
    
    # Filtro de data específica (tem prioridade sobre período rápido)
    if data_inicio or data_fim:
        try:
            if data_inicio:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                visitas = visitas.filter(data_entrada__date__gte=data_inicio_obj)
            if data_fim:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
                data_fim_completa = datetime.combine(data_fim_obj, time.max)
                visitas = visitas.filter(data_entrada__lte=data_fim_completa)
        except ValueError:
            pass
    elif periodo:
        hoje = timezone.localtime().date()
        if periodo == 'hoje':
            visitas = visitas.filter(data_entrada__date=hoje)
        elif periodo == 'semana':
            inicio_semana = hoje - timedelta(days=hoje.weekday())
            visitas = visitas.filter(data_entrada__date__gte=inicio_semana)
        elif periodo == 'mes':
            inicio_mes = hoje.replace(day=1)
            visitas = visitas.filter(data_entrada__date__gte=inicio_mes)
    
    if busca:
        visitas = visitas.filter(
            Q(visitante__nome_completo__icontains=busca) |
            Q(visitante__CPF__icontains=busca)
        )
    
    # Ordenação
    visitas = visitas.order_by('-data_entrada')
    
    # Paginação
    paginator = Paginator(visitas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_visitas = visitas.count()
    visitas_em_andamento = visitas.filter(data_saida__isnull=True).count()
    visitas_finalizadas = visitas.filter(data_saida__isnull=False).count()

    # Renderizar apenas a tabela e estatísticas
    table_html = render_to_string(
        'recepcao/includes/tabela_historico_visitas.html',
        {
            'page_obj': page_obj,
            'status_filtro': status,
            'periodo_filtro': periodo,
            'busca': busca,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
        },
        request=request
    )
    
    stats_html = render_to_string(
        'recepcao/includes/stats_historico_visitas.html',
        {
            'total_visitas': total_visitas,
            'visitas_em_andamento': visitas_em_andamento,
            'visitas_finalizadas': visitas_finalizadas,
        },
        request=request
    )

    return JsonResponse({
        'success': True,
        'html': table_html,
        'stats_html': stats_html,
        'total_visitas': total_visitas,
        'visitas_em_andamento': visitas_em_andamento,
        'visitas_finalizadas': visitas_finalizadas,
    })

# Carregamento em memória com TTL para evitar leituras repetidas do DB
known_face_encodings = []
known_face_ids = []
_known_loaded_at = None
_FACE_VECTORS_TTL_SECONDS = 600  # 10 minutos

def carregar_vetores_faciais():
    """Carrega vetores do banco e atualiza o cache em memória."""
    import time
    global known_face_encodings, known_face_ids, _known_loaded_at
    visitantes_com_vetor = Visitante.objects.filter(biometric_vector__isnull=False)
    known_face_encodings = [np.array(v.biometric_vector) for v in visitantes_com_vetor]
    known_face_ids = [v.id for v in visitantes_com_vetor]
    _known_loaded_at = time.time()
    if logger.isEnabledFor(logging.DEBUG):
        logger.debug("Carregados %s vetores faciais.", len(known_face_encodings))


def ensure_vetores_carregados():
    """Garante que o cache está válido; recarrega se vazio ou expirado."""
    import time
    global _known_loaded_at
    if not known_face_encodings or _known_loaded_at is None:
        carregar_vetores_faciais()
        return
    if (time.time() - _known_loaded_at) > _FACE_VECTORS_TTL_SECONDS:
        carregar_vetores_faciais()
        return

@csrf_exempt
@require_POST
def api_reconhecer_rosto(request):
    # Usa cache com TTL para evitar recargas a cada POST
    ensure_vetores_carregados()
    try:
        # 1. Lê o corpo da requisição JSON
        data = json.loads(request.body)
        image_data = data.get('image')

        if not image_data:
            return JsonResponse({'success': False, 'error': 'Nenhuma imagem enviada.'}, status=400)

        # 2. Decodifica a imagem base64
        # Remove o prefixo "data:image/jpeg;base64," se existir
        if 'base64,' in image_data:
            format, imgstr = image_data.split(';base64,')
            ext = format.split('/')[-1]
            image_data = ContentFile(base64.b64decode(imgstr), name=f'temp.{ext}')
        else:
            # Se não houver prefixo, assume que é base64 puro
            image_data = ContentFile(base64.b64decode(image_data), name='temp.jpg')

        # 3. Extrai o vetor da imagem
        unknown_embedding = get_face_embedding(image_data)

        if unknown_embedding is None:
            return JsonResponse({'success': False, 'error': 'Nenhum rosto detectado na imagem.'})
        
        if not known_face_encodings:
            return JsonResponse({'success': False, 'error': 'Nenhum visitante com dados faciais cadastrados.'})

        # 4. Calcula distâncias para todos os rostos conhecidos
        face_distances = face_recognition.face_distance(known_face_encodings, np.array(unknown_embedding))
        if len(face_distances) == 0:
            return JsonResponse({'success': False, 'error': 'Nenhum visitante com dados faciais cadastrados.'})

        # Melhor candidato (menor distância)
        best_index = int(np.argmin(face_distances))
        best_distance = float(face_distances[best_index])

        # Critério 1: distância máxima aceitável (mais rígido para reduzir falsos positivos)
        MAX_DISTANCE = 0.52
        if best_distance > MAX_DISTANCE:
            return JsonResponse({'success': False, 'error': 'Nenhum visitante correspondente encontrado.'})

        # Critério 2: diferença mínima para o segundo melhor (evita empates/vizinhos muito próximos)
        if len(face_distances) > 1:
            sorted_distances = sorted(face_distances)
            second_best_distance = float(sorted_distances[1])
            MIN_GAP = 0.12
            if (second_best_distance - best_distance) < MIN_GAP:
                return JsonResponse({'success': False, 'error': 'Nenhum visitante correspondente encontrado.'})

        visitante_id = known_face_ids[best_index]
        visitante = Visitante.objects.get(id=visitante_id)

        return JsonResponse({
            'success': True,
            'visitante_id': visitante.id, # Envia o ID para o redirect
            'nome_visitante': visitante.nome_completo
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Formato de requisição inválido (esperado JSON).'}, status=400)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro interno do servidor: {str(e)}'}, status=500)

@login_required(login_url='autenticacao:login_sistema')
@block_assessor
def totem_welcome(request):
    """
    Renderiza a página inicial de boas-vindas do totem.
    """
    return render(request, 'recepcao/totem_welcome.html')

@require_GET
def api_preload_face_vectors(request):
    """Pré-carrega os vetores faciais em memória para acelerar o reconhecimento."""
    try:
        import time
        ensure_vetores_carregados()
        age = None
        if _known_loaded_at:
            age = int(time.time() - _known_loaded_at)
        return JsonResponse({'success': True, 'count': len(known_face_encodings), 'age_seconds': age, 'ttl_seconds': _FACE_VECTORS_TTL_SECONDS})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def totem_finalize_search(request):
    """
    Renderiza a página de busca para finalizar uma visita.
    """
    return render(request, 'recepcao/totem_finalize_search.html')


def totem_confirmacao_identidade(request):
    """Tela intermediária para o visitante confirmar se é a pessoa reconhecida."""
    visitante_id = request.GET.get('visitante_id')
    if not visitante_id:
        messages.error(request, "ID do visitante não fornecido.")
        return redirect('recepcao:totem_identificacao')

    try:
        visitante = Visitante.objects.get(id=visitante_id)
    except Visitante.DoesNotExist:
        messages.error(request, "Visitante não encontrado.")
        return redirect('recepcao:totem_identificacao')

    foto_version = int(visitante.data_atualizacao.timestamp()) if visitante.data_atualizacao else int(timezone.now().timestamp())
    context = {
        'visitante': visitante,
        'visitante_foto_url': visitante.get_foto_url(size='large'),
        'visitante_foto_version': foto_version
    }
    return render(request, 'recepcao/totem_confirmacao_identidade.html', context)


def totem_destino(request):
    """
    Renderiza a página de seleção de destino para o visitante reconhecido.
    """
    visitante_id = request.GET.get('visitante_id')
    if not visitante_id:
        messages.error(request, "ID do visitante não fornecido.")
        return redirect('recepcao:totem_identificacao')

    try:
        visitante = Visitante.objects.get(id=visitante_id)
        foto_version = int(visitante.data_atualizacao.timestamp()) if visitante.data_atualizacao else int(timezone.now().timestamp())
        context = {
            'visitante': visitante,
            'visitante_foto_url': visitante.get_foto_url(size='large'),
            'visitante_foto_version': foto_version
        }
        return render(request, 'recepcao/totem_destino.html', context)
    except Visitante.DoesNotExist:
        messages.error(request, "Visitante não encontrado.")
        return redirect('recepcao:totem_identificacao')

def totem_identificacao(request):
    """
    Renderiza a página inicial do totem de autoatendimento.
    """
    return render(request, 'recepcao/totem_identificacao.html')

def totem_comprovante(request, visita_id):
    """
    Exibe a tela final com o comprovante da visita.
    Esta página pode ser recarregada.
    """
    visita = get_object_or_404(Visita, id=visita_id)
    context = {
        'visita': visita,
        'nome_exibicao': visita.visitante.nome_social if visita.visitante.nome_social else visita.visitante.nome_completo
    }
    return render(request, 'recepcao/totem_comprovante.html', context)

def totem_recadastro_facial(request):
    """Tela de fallback do totem: CPF + recadastro facial."""
    return render(request, 'recepcao/totem_recadastro_facial.html')

@require_GET
def api_get_setores(request):
    """
    API para retornar uma lista de setores (departamentos ou gabinetes) com todos os campos necessários para exibição de cards modernos.
    """
    tipo = request.GET.get('tipo')
    if tipo not in ['departamento', 'gabinete', 'gabinete_vereador']:
        return JsonResponse({'success': False, 'error': 'Tipo inválido'}, status=400)
    
    if tipo in ['gabinete', 'gabinete_vereador']:
        setores = Setor.objects.filter(tipo__in=['gabinete', 'gabinete_vereador'], ativo=True).order_by('nome_vereador')
    else:
        setores = Setor.objects.filter(tipo=tipo, ativo=True).order_by('nome_local')
    setores_data = []
    for setor in setores:
        dados = {
            'id': setor.id,
            'nome': setor.nome_vereador if tipo in ['gabinete', 'gabinete_vereador'] else setor.nome_local,
            'nome_vereador': setor.nome_vereador if tipo in ['gabinete', 'gabinete_vereador'] else None,
            'nome_local': setor.nome_local if tipo == 'departamento' else None,
            'tipo': setor.tipo,
            'localizacao': setor.get_localizacao_display(),
            'funcao': setor.get_funcao_display() if setor.funcao else None,
            'email': setor.email_vereador if tipo in ['gabinete', 'gabinete_vereador'] else setor.email,
            'foto_url': None, # Inicialmente nulo
            'aberto_agora': setor.esta_aberto()
        }
        
        # Prioridade: foto do gabinete (vereador), depois foto do assessor
        if tipo in ['gabinete', 'gabinete_vereador'] and setor.foto:
            dados['foto_url'] = setor.foto.url
        elif setor.usuario and hasattr(setor.usuario, 'assessor') and setor.usuario.assessor.foto:
            dados['foto_url'] = setor.usuario.assessor.foto.url

        setores_data.append(dados)

    return JsonResponse({'success': True, 'setores': setores_data})

@csrf_exempt
def api_registrar_visita_totem(request):

    if request.method != 'POST':
        logger.warning("api_registrar_visita_totem chamada com método GET.")
        return JsonResponse({'success': False, 'error': 'Método inválido'})

    logger.info("Iniciando o registro de visita via totem.")
    try:
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Decodificando JSON do corpo da requisição...")
        data = json.loads(request.body)
        visitante_id = data.get('visitante_id')
        setor_id = data.get('setor_id')
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Dados recebidos: visitante_id=%s, setor_id=%s", visitante_id, setor_id)

        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Buscando visitante com ID %s...", visitante_id)
        visitante = Visitante.objects.get(id=visitante_id)
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Visitante encontrado.")

        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Buscando setor com ID %s...", setor_id)
        setor = Setor.objects.get(id=setor_id)
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Setor encontrado.")

        visita_data_para_criar = {
            'visitante': visitante,
            'setor': setor,
            'localizacao': setor.localizacao,
            'status': 'em_andamento'
        }
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Dados para criação da visita: %s", visita_data_para_criar)

        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Criando registro da visita no banco de dados...")
        nova_visita = Visita.objects.create(**visita_data_para_criar)
        logger.info("Visita criada com sucesso: id=%s", nova_visita.id)

        if logger.isEnabledFor(logging.DEBUG):
            logger.debug("Enviando ID da nova visita para o frontend: %s", nova_visita.id)
        return JsonResponse({'success': True, 'visita_id': nova_visita.id})

    except (Visitante.DoesNotExist, Setor.DoesNotExist) as e:
        logger.error(f"Erro de objeto não encontrado: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Visitante ou Setor não encontrado.'}, status=404)
    except json.JSONDecodeError:
        logger.error("Erro ao decodificar JSON da requisição.")
        return JsonResponse({'success': False, 'error': 'Requisição mal formatada.'}, status=400)
    except Exception as e:
        logger.exception("Erro inesperado e fatal em api_registrar_visita_totem:")
        return JsonResponse({'success': False, 'error': 'Ocorreu uma falha interna no servidor.'}, status=500)


@csrf_exempt
@require_POST
def api_validar_cpf_totem(request):
    """Valida CPF digitado no totem e retorna dados básicos do visitante."""
    try:
        data = json.loads(request.body)
        cpf = data.get('cpf', '').strip()

        if not cpf:
            return JsonResponse({'success': False, 'error': 'CPF não informado.'}, status=400)

        try:
            visitante = Visitante.objects.get(CPF=cpf)
        except Visitante.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'CPF_NOT_FOUND'}, status=404)

        return JsonResponse({
            'success': True,
            'visitante_id': visitante.id,
            'nome_visitante': visitante.nome_completo,
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Requisição mal formatada.'}, status=400)
    except Exception as e:
        logger.exception("Erro em api_validar_cpf_totem:")
        return JsonResponse({'success': False, 'error': 'Falha interna do servidor.'}, status=500)


@csrf_exempt
@require_POST
def api_re_enroll_face(request):
    """Atualiza o vetor biométrico de um visitante a partir de uma nova captura no totem."""
    try:
        data = json.loads(request.body)
        visitante_id = data.get('visitante_id')
        image_data = data.get('image')

        if not visitante_id or not image_data:
            return JsonResponse({'success': False, 'error': 'Dados insuficientes.'}, status=400)

        try:
            visitante = Visitante.objects.get(id=visitante_id)
        except Visitante.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Visitante não encontrado.'}, status=404)

        # Decodificar imagem base64 (mesma lógica de api_reconhecer_rosto)
        if 'base64,' in image_data:
            format, imgstr = image_data.split(';base64,')
            ext = format.split('/')[-1]
            image_binary = base64.b64decode(imgstr)
            image_file = ContentFile(image_binary, name=f'temp.{ext}')
        else:
            image_binary = base64.b64decode(image_data)
            image_file = ContentFile(image_binary, name='temp.jpg')

        # Verificar tamanho do rosto na imagem para evitar capturas muito distantes
        try:
            import io
            image_array = face_recognition.load_image_file(io.BytesIO(image_binary))
            face_locations = face_recognition.face_locations(image_array)
            if not face_locations:
                return JsonResponse({'success': False, 'error': 'Nenhum rosto detectado na imagem.'})

            top, right, bottom, left = face_locations[0]
            face_width = right - left
            image_width = image_array.shape[1]
            face_width_ratio = face_width / image_width if image_width else 0

            MIN_FACE_WIDTH_RATIO = 0.20  # rosto deve ocupar pelo menos ~20% da largura
            if face_width_ratio < MIN_FACE_WIDTH_RATIO:
                return JsonResponse({'success': False, 'error': 'Rosto muito distante da câmera. Aproxime-se e tente novamente.'})
        except Exception as e:
            logger.warning(f"Falha ao calcular tamanho do rosto na recaptura: {e}")

        # Gerar novo vetor biométrico para validar a imagem
        embedding = get_face_embedding(image_file)
        if embedding is None:
            return JsonResponse({'success': False, 'error': 'Nenhum rosto detectado na imagem.'})

        # Substituir a foto do visitante pela nova captura
        # O save() do modelo Visitante já trata geração de thumbnails e
        # recalcula o vetor biométrico a partir de self.foto.
        visitante.foto = image_file
        visitante.save()

        # Limpa o cache das URLs da foto para exibir imediatamente a nova imagem
        cache.delete(f'visitante_foto_{visitante.id}_thumbnail')
        cache.delete(f'visitante_foto_{visitante.id}_medium')
        cache.delete(f'visitante_foto_{visitante.id}_large')
        cache.delete(f'visitante_foto_{visitante.id}_original')

        return JsonResponse({
            'success': True,
            'visitante_id': visitante.id,
            'nome_visitante': visitante.nome_completo,
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Requisição mal formatada.'}, status=400)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        logger.exception("Erro em api_re_enroll_face:")
        return JsonResponse({'success': False, 'error': 'Falha interna do servidor.'}, status=500)

@csrf_exempt
@require_POST
def upload_foto_webcam(request):
    """View específica para upload de foto via webcam"""
    import logging
    import base64
    import uuid
    from django.core.files.base import ContentFile
    from .utils.image_utils import process_image
    
    # usa logger do módulo
    
    try:
        # Receber dados da imagem em base64
        image_data = request.POST.get('image_data')
        if not image_data:
            return JsonResponse({'success': False, 'error': 'Dados da imagem não fornecidos'})
        
        # Remover o prefixo "data:image/jpeg;base64,"
        if 'base64,' in image_data:
            image_data = image_data.split('base64,')[1]
        
        # Decodificar base64
        image_binary = base64.b64decode(image_data)
        
        # Criar arquivo temporário
        filename = f"webcam_{uuid.uuid4().hex}.jpg"
        image_file = ContentFile(image_binary, name=filename)
        
        # Gerar vetor biométrico da foto
        try:
            embedding = get_face_embedding(image_file)
            if embedding is None:
                return JsonResponse({
                    'success': False, 
                    'error': 'Nenhum rosto detectado na imagem. Tire uma nova foto.'
                })
        except ValueError as e:
            return JsonResponse({
                'success': False, 
                'error': str(e)
            })
        except Exception as e:
            logger.error(f"Erro ao gerar vetor biométrico: {str(e)}")
            embedding = None
        
        # Processar imagem
        processed_images = process_image(image_file)
        
        # Salvar as versões processadas fisicamente
        media_root = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'media', 'fotos_visitantes')
        os.makedirs(media_root, exist_ok=True)
        
        # Salvar cada versão
        saved_paths = {}
        for size, img_file in processed_images.items():
            file_path = os.path.join(media_root, img_file.name)
            with open(file_path, 'wb') as f:
                f.write(img_file.read())
            saved_paths[size] = f"fotos_visitantes/{img_file.name}"
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Arquivo salvo: %s", file_path)
        
        # Também salvar o original
        original_path = os.path.join(media_root, filename)
        with open(original_path, 'wb') as f:
            f.write(image_binary)
        saved_paths['original'] = f"fotos_visitantes/{filename}"
        
        return JsonResponse({
            'success': True, 
            'paths': saved_paths,
            'biometric_vector': embedding,
            'message': 'Foto processada e salva com sucesso'
        })
        
    except Exception as e:
        logger.error(f"Erro ao processar foto da webcam: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@require_GET
def api_buscar_visitante_ativo(request):
    query = request.GET.get('query', '').strip()
    if len(query) < 3:
        return JsonResponse({'success': False, 'error': 'Forneça ao menos 3 caracteres para a busca.'}, status=400)

    # Passo 1: Encontrar o visitante por nome ou CPF, sem filtrar por visita.
    visitantes_qs = Visitante.objects.filter(
        Q(nome_completo__icontains=query) | Q(CPF__iexact=query)
    ).distinct()

    if not visitantes_qs.exists():
        return JsonResponse({'success': False, 'error': 'Nenhum visitante encontrado com os dados informados.'})

    # Passo 2: Preparar os dados, verificando as visitas para cada um.
    visitantes_data = []
    for visitante in visitantes_qs:
        # Agora sim, filtramos as visitas em andamento para este visitante específico.
        visitas_em_andamento = visitante.visita_set.filter(status='em_andamento')
        visitas_data_list = [{
            'id': v.id,
            'setor': v.setor.nome_vereador if v.setor.tipo in ['gabinete', 'gabinete_vereador'] else v.setor.nome_local,
            'data_entrada': v.data_entrada.strftime('%d/%m/%Y %H:%M')
        } for v in visitas_em_andamento]

        # Inclui o visitante na resposta, mesmo que não tenha visitas ativas.
        visitantes_data.append({
            'id': visitante.id,
            'nome': visitante.nome_completo,
            'visitas': visitas_data_list
        })

    return JsonResponse({'success': True, 'visitantes': visitantes_data})

@csrf_exempt
@require_POST
def api_finalizar_visitas(request):
    try:
        data = json.loads(request.body)
        visitante_id = data.get('visitante_id')
        if not visitante_id:
            return JsonResponse({'success': False, 'error': 'ID do visitante não fornecido.'}, status=400)

        visitas_finalizadas = Visita.objects.filter(
            visitante_id=visitante_id,
            status='em_andamento'
        ).update(
            status='finalizada',
            data_saida=timezone.now()
        )

        if visitas_finalizadas == 0:
            return JsonResponse({'success': False, 'error': 'Nenhuma visita em andamento para finalizar.'})

        return JsonResponse({'success': True, 'message': f'{visitas_finalizadas} visita(s) finalizada(s) com sucesso.'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Requisição mal formatada.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Ocorreu uma falha interna: {str(e)}'}, status=500)

@login_required(login_url='autenticacao:login_sistema')
@require_POST
def finalizar_multiplas_visitas(request):
    """
    Finaliza múltiplas visitas de uma só vez.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # Log dos dados recebidos para debugging
    logger.info(f"POST data recebido: {request.POST}")
    logger.info(f"Content-Type: {request.content_type}")
    logger.info(f"Method: {request.method}")
    
    visitas_ids = request.POST.getlist('visitas_ids')
    logger.info(f"IDs de visitas extraídos: {visitas_ids}")
    
    if not visitas_ids:
        logger.warning("Nenhuma visita selecionada")
        return JsonResponse({'success': False, 'error': 'Nenhuma visita selecionada.'}, status=400)
    
    try:
        # Verificar se todas as visitas existem e estão em andamento
        visitas = Visita.objects.filter(
            id__in=visitas_ids,
            status='em_andamento',
            data_saida__isnull=True
        )
        
        logger.info(f"Visitas encontradas: {visitas.count()}")
        
        if not visitas.exists():
            logger.warning("Nenhuma visita válida encontrada")
            return JsonResponse({'success': False, 'error': 'Nenhuma visita válida encontrada para finalizar.'})
        
        # Finalizar as visitas
        visitas_finalizadas = visitas.update(
            status='finalizada',
            data_saida=timezone.now()
        )
        
        logger.info(f"Visitas finalizadas com sucesso: {visitas_finalizadas}")
        
        return JsonResponse({
            'success': True,
            'finalizadas': visitas_finalizadas,
            'message': f'{visitas_finalizadas} visita(s) finalizada(s) com sucesso!'
        })
        
    except Exception as e:
        logger.error(f"Erro ao finalizar visitas: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro ao finalizar visitas: {str(e)}'}, status=500)

